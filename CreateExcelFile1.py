import openpyxl

wb = openpyxl.Workbook()
#wb.sheetnames
sheet = wb.active
sheet.title
sheet.title = "NewSheet1"
#wb.sheetnames
wb.create_sheet()
print("a new sheet has been created")
wb.create_sheet(index = 2, title = "NewSheet2")
print("NewSheet2 has been created")
del wb["Sheet"]
print("Sheet has been deleted")
sheet = wb["NewSheet2"]
sheet["A1"] = "Hello, World!"
topleftcell = sheet["A1"]
print("Cell %s is %s" % (topleftcell.coordinate, topleftcell.value))
wb.save("CreateExcelFile.xlsx")